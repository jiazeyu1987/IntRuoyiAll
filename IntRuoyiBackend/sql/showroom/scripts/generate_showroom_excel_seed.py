from __future__ import annotations

import argparse
import hashlib
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


EXPECTED_HEADERS = [
    "展厅ID",
    "展厅名称",
    "排序",
    "产品ID",
    "产品名称",
    "英文名称",
    "产品描述",
    "注册证名称",
    "注册证号",
    "生效时间",
    "所属公司",
    "产品来源",
]
SEED_AUTHOR = "showroom-seed"
SEED_TIMESTAMP = "2026-05-19 00:00:00"
MAIN_COMPANY_NAME = "瑛泰"


@dataclass(frozen=True)
class HallText:
    name_en: str
    description: str
    description_en: str


HALL_TEXT_BY_CODE = {
    "hall_01": HallText(
        "Cardiac Intervention Implant Showcase",
        "集中展示心内介植入相关产品，覆盖冠脉介入、通路建立及术中辅助器械。",
        "Presents cardiac interventional implant products, covering coronary intervention, access establishment, and procedure-support devices.",
    ),
    "hall_02": HallText(
        "Cardiac Implant Showcase",
        "集中展示心脏植入及相关介入产品，覆盖结构性心脏病、心脏通路及配套器械。",
        "Presents cardiac implant and related interventional products, covering structural heart disease, cardiac access, and supporting devices.",
    ),
    "hall_03": HallText(
        "Peripheral Intervention Implant Showcase",
        "集中展示外周介植入相关产品，覆盖主动脉、外周血管及分支血管治疗器械。",
        "Presents peripheral interventional implant products, covering aortic, peripheral vascular, and branch-vessel treatment devices.",
    ),
    "hall_04": HallText(
        "Neuro Intervention Implant Showcase",
        "集中展示神经介植入相关产品，覆盖颅内血管通路、取栓、输送及支撑器械。",
        "Presents neuro interventional implant products, covering intracranial vascular access, thrombectomy, delivery, and support devices.",
    ),
    "hall_05": HallText(
        "Exosome and Focused Ultrasound Showcase",
        "集中展示外泌体应用与聚焦超声相关产品，覆盖无创透皮、能量治疗及配套解决方案。",
        "Presents exosome application and focused ultrasound products, covering non-invasive transdermal delivery, energy therapy, and supporting solutions.",
    ),
    "hall_06": HallText(
        "Orthopedics and Urology Products Showcase",
        "集中展示骨科与泌尿方向产品，覆盖关节介入、骨科手术及泌尿治疗相关器械。",
        "Presents orthopedics and urology products, covering joint intervention, orthopedic procedures, and urological treatment devices.",
    ),
    "hall_07": HallText(
        "Non-interventional Products Showcase",
        "集中展示非介入类医疗与健康产品，覆盖材料、消费医疗及配套健康管理方案。",
        "Presents non-interventional medical and health products, covering materials, consumer medical products, and supporting health-management solutions.",
    ),
    "hall_08": HallText(
        "Medical Standard Components Showcase",
        "集中展示医疗器械标准件与基础组件，覆盖导管、连接件、耗材组件及制造配套。",
        "Presents standard medical device components and foundational parts, covering catheters, connectors, consumable components, and manufacturing support.",
    ),
}


@dataclass(frozen=True)
class ProductRow:
    hall_code: str
    hall_name: str
    hall_product_order: int
    product_code: str
    name_cn: str
    name_en: str
    indication_content: str | None
    registration_certificate: str | None
    owner_company_name: str | None
    owner_company_id: int | None
    product_owner_type: str | None
    lifecycle_stage: str


@dataclass(frozen=True)
class CompanySeed:
    id: int
    company_code: str
    display_name: str
    company_type: str


@dataclass(frozen=True)
class HallSeed:
    id: int
    hall_code: str
    hall_name: str
    hall_name_en: str
    description: str
    description_en: str
    display_order: int


@dataclass(frozen=True)
class ProductSeed:
    id: int
    product_code: str
    revision_id: int
    row: ProductRow


def repo_root() -> Path:
    return Path(__file__).resolve().parents[3]


def default_input_path() -> Path:
    return repo_root().parent / "resource" / "展厅产品与描述清单.xlsx"


def default_output_path() -> Path:
    return repo_root() / "sql" / "showroom" / "20260519_showroom_excel_seed.sql"


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return f"{value.year}.{value.month}.{value.day}"
    return str(value).strip()


def optional_text(value: object) -> str | None:
    text = normalize_text(value)
    return text if text else None


def normalize_hall_name(value: object) -> str:
    text = normalize_text(value)
    if text.endswith("展厅"):
        return text[:-len("展厅")] + "展柜"
    return text


def sql_literal(value: object) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "b'1'" if value else "b'0'"
    if isinstance(value, int):
        return str(value)
    text = str(value).replace("'", "''")
    return f"'{text}'"


def registration_certificate_text(name: str | None, number: str | None, effective_date: str | None) -> str | None:
    lines: list[str] = []
    if name:
        lines.append(f"注册证名称：{name}")
    if number:
        lines.append(f"注册证号：{number}")
    if effective_date:
        lines.append(f"生效时间：{effective_date}")
    return "\n".join(lines) if lines else None


def company_code(name: str) -> str:
    return "OWNER_" + hashlib.md5(name.encode("utf-8")).hexdigest()[:12].upper()


def parse_workbook(path: Path) -> tuple[list[CompanySeed], list[HallSeed], list[ProductSeed], list[tuple[str, str]]]:
    workbook = load_workbook(path, data_only=False)
    if len(workbook.worksheets) != 2:
        raise ValueError(f"SHOWROOM_EXCEL_INVALID: expected 2 sheets, got {len(workbook.worksheets)}")

    detail_sheet = workbook.worksheets[1]
    headers = [normalize_text(cell.value) for cell in detail_sheet[1]]
    if headers != EXPECTED_HEADERS:
        raise ValueError(f"SHOWROOM_EXCEL_INVALID: unexpected headers {headers}")

    hall_meta: dict[str, HallSeed] = {}
    hall_sorts: dict[str, set[int]] = {}
    companies_by_name: dict[str, CompanySeed] = {}
    products_by_code: dict[str, ProductSeed] = {}
    empty_descriptions: list[tuple[str, str]] = []

    hall_sequence = 1
    company_sequence = 1
    product_sequence = 1

    for row_index, row in enumerate(detail_sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        hall_code = normalize_text(row[0])
        hall_name = normalize_hall_name(row[1])
        hall_product_order_text = normalize_text(row[2])
        product_code = normalize_text(row[3])
        name_cn = normalize_text(row[4])
        name_en = normalize_text(row[5])
        indication_content = optional_text(row[6])
        registration_name = optional_text(row[7])
        registration_number = optional_text(row[8])
        effective_date = optional_text(row[9])
        owner_company_name = optional_text(row[10])

        if not hall_code or not hall_name or not hall_product_order_text or not product_code or not name_cn:
            raise ValueError(f"SHOWROOM_EXCEL_INVALID: required column missing at row {row_index}")

        try:
            hall_product_order = int(float(hall_product_order_text))
        except ValueError as exc:
            raise ValueError(
                f"SHOWROOM_EXCEL_INVALID: hall display order is not numeric at row {row_index}"
            ) from exc

        existing_hall = hall_meta.get(hall_code)
        if existing_hall is None:
            hall_text = HALL_TEXT_BY_CODE.get(hall_code)
            if hall_text is None:
                raise ValueError(f"SHOWROOM_EXCEL_INVALID: hall code {hall_code} has no bilingual description mapping")
            hall_meta[hall_code] = HallSeed(
                id=hall_sequence,
                hall_code=hall_code,
                hall_name=hall_name,
                hall_name_en=hall_text.name_en,
                description=hall_text.description,
                description_en=hall_text.description_en,
                display_order=hall_sequence,
            )
            hall_sorts[hall_code] = {hall_product_order}
            hall_sequence += 1
        else:
            if existing_hall.hall_name != hall_name:
                raise ValueError(
                    f"SHOWROOM_EXCEL_INVALID: hall code {hall_code} maps to multiple names"
                )
            if hall_product_order in hall_sorts[hall_code]:
                raise ValueError(
                    f"SHOWROOM_EXCEL_INVALID: duplicate display order {hall_product_order} in hall {hall_code}"
                )
            hall_sorts[hall_code].add(hall_product_order)

        owner_company_id = None
        product_owner_type = None
        if owner_company_name:
            if owner_company_name not in companies_by_name:
                companies_by_name[owner_company_name] = CompanySeed(
                    id=company_sequence,
                    company_code=company_code(owner_company_name),
                    display_name=owner_company_name,
                    company_type="MAIN" if owner_company_name == MAIN_COMPANY_NAME else "SUBSIDIARY",
                )
                company_sequence += 1
            owner_company_id = companies_by_name[owner_company_name].id
            product_owner_type = "YINGTAI" if owner_company_name == MAIN_COMPANY_NAME else "SUBSIDIARY"

        if product_code in products_by_code:
            raise ValueError(f"SHOWROOM_EXCEL_INVALID: duplicate product code {product_code}")

        registration_certificate = registration_certificate_text(
            registration_name, registration_number, effective_date
        )
        lifecycle_stage = "REGISTERED" if registration_number or effective_date else "R_AND_D"
        parsed_row = ProductRow(
            hall_code=hall_code,
            hall_name=hall_name,
            hall_product_order=hall_product_order,
            product_code=product_code,
            name_cn=name_cn,
            name_en=name_en,
            indication_content=indication_content,
            registration_certificate=registration_certificate,
            owner_company_name=owner_company_name,
            owner_company_id=owner_company_id,
            product_owner_type=product_owner_type,
            lifecycle_stage=lifecycle_stage,
        )
        products_by_code[product_code] = ProductSeed(
            id=product_sequence,
            product_code=product_code,
            revision_id=1000 + product_sequence,
            row=parsed_row,
        )
        product_sequence += 1

        if indication_content is None:
            empty_descriptions.append((product_code, name_cn))

    return (
        list(companies_by_name.values()),
        list(hall_meta.values()),
        list(products_by_code.values()),
        empty_descriptions,
    )


def render_insert(table_name: str, columns: list[str], rows: Iterable[tuple[object, ...]]) -> str:
    rendered_rows = []
    for row in rows:
        rendered_rows.append("  (" + ", ".join(sql_literal(value) for value in row) + ")")
    return (
        f"INSERT INTO {table_name} ({', '.join(columns)})\nVALUES\n"
        + ",\n".join(rendered_rows)
        + ";\n"
    )


def render_sql(companies: list[CompanySeed], halls: list[HallSeed], products: list[ProductSeed],
               empty_descriptions: list[tuple[str, str]], workbook_path: Path) -> str:
    hall_lookup = {hall.hall_code: hall for hall in halls}
    mapping_rows = [
        (
            2000 + index,
            hall_lookup[product.row.hall_code].id,
            product.id,
            product.row.hall_product_order,
            SEED_AUTHOR,
            SEED_TIMESTAMP,
            SEED_AUTHOR,
            SEED_TIMESTAMP,
            False,
            0,
        )
        for index, product in enumerate(products, start=1)
    ]

    company_rows = [
        (
            company.id,
            company.company_code,
            company.display_name,
            company.company_type,
            None,
            0,
            False,
            "DRAFT_ONLY",
            SEED_AUTHOR,
            SEED_TIMESTAMP,
            SEED_AUTHOR,
            SEED_TIMESTAMP,
            False,
            0,
        )
        for company in companies
    ]

    hall_rows = [
        (
            hall.id,
            hall.hall_code,
            hall.hall_name,
            hall.hall_name_en,
            hall.description,
            hall.description_en,
            hall.display_order,
            "ACTIVE",
            SEED_AUTHOR,
            SEED_TIMESTAMP,
            SEED_AUTHOR,
            SEED_TIMESTAMP,
            False,
            0,
        )
        for hall in halls
    ]

    product_rows = [
        (
            product.id,
            product.product_code,
            None,
            1,
            not bool(product.row.name_cn and product.row.name_en),
            "DRAFT_ONLY",
            SEED_AUTHOR,
            SEED_TIMESTAMP,
            SEED_AUTHOR,
            SEED_TIMESTAMP,
            False,
            0,
        )
        for product in products
    ]

    product_revision_rows = [
        (
            product.revision_id,
            product.id,
            1,
            "DRAFT",
            product.row.name_cn,
            product.row.name_en,
            product.row.owner_company_id,
            product.row.product_owner_type,
            product.row.lifecycle_stage,
            None,
            None,
            product.row.registration_certificate,
            product.row.indication_content,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            SEED_AUTHOR,
            SEED_TIMESTAMP,
            SEED_AUTHOR,
            SEED_TIMESTAMP,
            False,
            0,
        )
        for product in products
    ]

    empty_description_text = ", ".join(f"{code}:{name}" for code, name in empty_descriptions)
    lines = [
        "-- Showroom Excel seed generated from 展厅产品与描述清单.xlsx",
        f"-- Source workbook: {workbook_path.as_posix()}",
        f"-- Hall count: {len(halls)}",
        f"-- Product count: {len(products)}",
        f"-- Non-empty company count: {len(companies)}",
        f"-- Empty description products: {empty_description_text}",
        "",
        "DROP PROCEDURE IF EXISTS init_showroom_excel_seed;",
        "DELIMITER //",
        "CREATE PROCEDURE init_showroom_excel_seed()",
        "BEGIN",
        "  DECLARE v_live_products BIGINT DEFAULT 0;",
        "  DECLARE v_revision_advanced BIGINT DEFAULT 0;",
        "  DECLARE v_company_revision_rows BIGINT DEFAULT 0;",
        "  DECLARE v_relation_rows BIGINT DEFAULT 0;",
        "  DECLARE v_change_request_rows BIGINT DEFAULT 0;",
        "  DECLARE v_change_request_item_rows BIGINT DEFAULT 0;",
        "  DECLARE v_assignment_rows BIGINT DEFAULT 0;",
        "  DECLARE v_comment_rows BIGINT DEFAULT 0;",
        "  DECLARE v_narration_rows BIGINT DEFAULT 0;",
        "  DECLARE v_preview_rows BIGINT DEFAULT 0;",
        "  DECLARE v_company_diff BIGINT DEFAULT 0;",
        "  DECLARE v_company_rows BIGINT DEFAULT 0;",
        "  DECLARE v_product_diff BIGINT DEFAULT 0;",
        "  DECLARE v_product_rows BIGINT DEFAULT 0;",
        "  DECLARE v_product_revision_diff BIGINT DEFAULT 0;",
        "  DECLARE v_product_revision_rows BIGINT DEFAULT 0;",
        "  DECLARE v_hall_diff BIGINT DEFAULT 0;",
        "  DECLARE v_hall_rows BIGINT DEFAULT 0;",
        "  DECLARE v_hall_product_diff BIGINT DEFAULT 0;",
        "  DECLARE v_hall_product_rows BIGINT DEFAULT 0;",
        "  DECLARE v_seed_tenant_id BIGINT DEFAULT 1;",
        "",
        "  CREATE TEMPORARY TABLE tmp_showroom_seed_company LIKE showroom_company;",
        "  CREATE TEMPORARY TABLE tmp_showroom_seed_product LIKE showroom_product;",
        "  CREATE TEMPORARY TABLE tmp_showroom_seed_product_revision LIKE showroom_product_revision;",
        "  CREATE TEMPORARY TABLE tmp_showroom_seed_hall LIKE showroom_hall;",
        "  CREATE TEMPORARY TABLE tmp_showroom_seed_hall_product LIKE showroom_hall_product;",
        "",
        render_insert(
            "tmp_showroom_seed_company",
            [
                "id", "company_code", "display_name", "company_type", "current_revision_id",
                "current_revision_no", "incomplete_flag", "status", "creator", "create_time",
                "updater", "update_time", "deleted", "tenant_id",
            ],
            company_rows,
        ).rstrip(),
        "",
        render_insert(
            "tmp_showroom_seed_hall",
            [
                "id", "hall_code", "name", "name_en", "description", "description_en",
                "display_order", "status", "creator", "create_time", "updater", "update_time",
                "deleted", "tenant_id",
            ],
            hall_rows,
        ).rstrip(),
        "",
        render_insert(
            "tmp_showroom_seed_product",
            [
                "id", "product_code", "current_revision_id", "current_revision_no", "incomplete_flag",
                "status", "creator", "create_time", "updater", "update_time", "deleted", "tenant_id",
            ],
            product_rows,
        ).rstrip(),
        "",
        render_insert(
            "tmp_showroom_seed_product_revision",
            [
                "id", "product_id", "revision_no", "status", "name_cn", "name_en", "owner_company_id",
                "product_owner_type", "lifecycle_stage", "target_market", "pipeline_layout",
                "registration_certificate", "indication_content", "core_selling_points",
                "model_specification", "clinical_effect", "fim_status", "submitted_by", "approved_by",
                "published_at", "creator", "create_time", "updater", "update_time", "deleted", "tenant_id",
            ],
            product_revision_rows,
        ).rstrip(),
        "",
        render_insert(
            "tmp_showroom_seed_hall_product",
            [
                "id", "hall_id", "product_id", "display_order", "creator", "create_time", "updater",
                "update_time", "deleted", "tenant_id",
            ],
            mapping_rows,
        ).rstrip(),
        "",
        "  UPDATE tmp_showroom_seed_company SET tenant_id = v_seed_tenant_id;",
        "  UPDATE tmp_showroom_seed_product SET tenant_id = v_seed_tenant_id;",
        "  UPDATE tmp_showroom_seed_product_revision SET tenant_id = v_seed_tenant_id;",
        "  UPDATE tmp_showroom_seed_hall SET tenant_id = v_seed_tenant_id;",
        "  UPDATE tmp_showroom_seed_hall_product SET tenant_id = v_seed_tenant_id;",
        "",
        "  SELECT COUNT(*) INTO v_live_products",
        "  FROM showroom_product",
        "  WHERE tenant_id = v_seed_tenant_id AND deleted = b'0' AND current_revision_id IS NOT NULL;",
        "  IF v_live_products > 0 THEN",
        "    SIGNAL SQLSTATE '45000' SET MESSAGE_TEXT = 'SHOWROOM_SEED_BLOCKED: live product revisions already exist';",
        "  END IF;",
        "",
        "  SELECT COUNT(*) INTO v_revision_advanced",
        "  FROM showroom_product",
        "  WHERE tenant_id = v_seed_tenant_id AND deleted = b'0' AND current_revision_no > 1;",
        "  IF v_revision_advanced > 0 THEN",
        "    SIGNAL SQLSTATE '45000' SET MESSAGE_TEXT = 'SHOWROOM_SEED_BLOCKED: product revisions already advanced beyond initialization';",
        "  END IF;",
        "",
        "  SELECT COUNT(*) INTO v_company_revision_rows FROM showroom_company_revision",
        "  WHERE tenant_id = v_seed_tenant_id AND deleted = b'0';",
        "  IF v_company_revision_rows > 0 THEN",
        "    SIGNAL SQLSTATE '45000' SET MESSAGE_TEXT = 'SHOWROOM_SEED_BLOCKED: company revisions already exist';",
        "  END IF;",
        "",
        "  SELECT COUNT(*) INTO v_relation_rows FROM showroom_product_revision_relation",
        "  WHERE tenant_id = v_seed_tenant_id AND deleted = b'0';",
        "  IF v_relation_rows > 0 THEN",
        "    SIGNAL SQLSTATE '45000' SET MESSAGE_TEXT = 'SHOWROOM_SEED_BLOCKED: product relations already exist';",
        "  END IF;",
        "",
        "  SELECT COUNT(*) INTO v_change_request_rows FROM showroom_change_request",
        "  WHERE tenant_id = v_seed_tenant_id AND deleted = b'0';",
        "  IF v_change_request_rows > 0 THEN",
        "    SIGNAL SQLSTATE '45000' SET MESSAGE_TEXT = 'SHOWROOM_SEED_BLOCKED: change requests already exist';",
        "  END IF;",
        "",
        "  SELECT COUNT(*) INTO v_change_request_item_rows FROM showroom_change_request_item",
        "  WHERE tenant_id = v_seed_tenant_id AND deleted = b'0';",
        "  IF v_change_request_item_rows > 0 THEN",
        "    SIGNAL SQLSTATE '45000' SET MESSAGE_TEXT = 'SHOWROOM_SEED_BLOCKED: change request items already exist';",
        "  END IF;",
        "",
        "  SELECT COUNT(*) INTO v_assignment_rows FROM showroom_field_assignment",
        "  WHERE tenant_id = v_seed_tenant_id AND deleted = b'0';",
        "  IF v_assignment_rows > 0 THEN",
        "    SIGNAL SQLSTATE '45000' SET MESSAGE_TEXT = 'SHOWROOM_SEED_BLOCKED: assignments already exist';",
        "  END IF;",
        "",
        "  SELECT COUNT(*) INTO v_comment_rows FROM showroom_product_comment",
        "  WHERE tenant_id = v_seed_tenant_id AND deleted = b'0';",
        "  IF v_comment_rows > 0 THEN",
        "    SIGNAL SQLSTATE '45000' SET MESSAGE_TEXT = 'SHOWROOM_SEED_BLOCKED: product comments already exist';",
        "  END IF;",
        "",
        "  SELECT COUNT(*) INTO v_narration_rows FROM showroom_narration_version",
        "  WHERE tenant_id = v_seed_tenant_id AND deleted = b'0';",
        "  IF v_narration_rows > 0 THEN",
        "    SIGNAL SQLSTATE '45000' SET MESSAGE_TEXT = 'SHOWROOM_SEED_BLOCKED: narration versions already exist';",
        "  END IF;",
        "",
        "  SELECT COUNT(*) INTO v_preview_rows FROM showroom_preview_asset_version",
        "  WHERE tenant_id = v_seed_tenant_id AND deleted = b'0';",
        "  IF v_preview_rows > 0 THEN",
        "    SIGNAL SQLSTATE '45000' SET MESSAGE_TEXT = 'SHOWROOM_SEED_BLOCKED: preview asset versions already exist';",
        "  END IF;",
        "",
        "  SELECT COUNT(*) INTO v_company_rows FROM showroom_company",
        "  WHERE tenant_id = v_seed_tenant_id AND deleted = b'0';",
        "  IF v_company_rows > 0 THEN",
        "    SELECT COUNT(*) INTO v_company_diff FROM (",
        "    SELECT c.id",
        "    FROM showroom_company c",
        "    LEFT JOIN tmp_showroom_seed_company s",
        "      ON c.id = s.id",
        "     AND c.company_code = s.company_code",
        "     AND c.display_name = s.display_name",
        "     AND c.company_type = s.company_type",
        "     AND IFNULL(c.current_revision_id, 0) = IFNULL(s.current_revision_id, 0)",
        "     AND IFNULL(c.current_revision_no, 0) = IFNULL(s.current_revision_no, 0)",
        "     AND c.incomplete_flag = s.incomplete_flag",
        "     AND c.status = s.status",
        "     AND c.tenant_id = s.tenant_id",
        "    WHERE c.tenant_id = v_seed_tenant_id AND c.deleted = b'0' AND s.id IS NULL",
        "    UNION ALL",
        "    SELECT s.id",
        "    FROM tmp_showroom_seed_company s",
        "    LEFT JOIN showroom_company c",
        "      ON c.id = s.id",
        "     AND c.deleted = b'0'",
        "     AND c.tenant_id = s.tenant_id",
        "    WHERE c.id IS NULL",
        "    ) diff_company;",
        "    IF v_company_diff > 0 THEN",
        "      SIGNAL SQLSTATE '45000' SET MESSAGE_TEXT = 'SHOWROOM_SEED_BLOCKED: company master data no longer matches the initialization seed';",
        "    END IF;",
        "  END IF;",
        "",
        "  SELECT COUNT(*) INTO v_product_rows FROM showroom_product",
        "  WHERE tenant_id = v_seed_tenant_id AND deleted = b'0';",
        "  IF v_product_rows > 0 THEN",
        "    SELECT COUNT(*) INTO v_product_diff FROM (",
        "    SELECT p.id",
        "    FROM showroom_product p",
        "    LEFT JOIN tmp_showroom_seed_product s",
        "      ON p.id = s.id",
        "     AND p.product_code = s.product_code",
        "     AND IFNULL(p.current_revision_id, 0) = IFNULL(s.current_revision_id, 0)",
        "     AND IFNULL(p.current_revision_no, 0) = IFNULL(s.current_revision_no, 0)",
        "     AND p.incomplete_flag = s.incomplete_flag",
        "     AND p.status = s.status",
        "     AND p.tenant_id = s.tenant_id",
        "    WHERE p.tenant_id = v_seed_tenant_id AND p.deleted = b'0' AND s.id IS NULL",
        "    UNION ALL",
        "    SELECT s.id",
        "    FROM tmp_showroom_seed_product s",
        "    LEFT JOIN showroom_product p",
        "      ON p.id = s.id",
        "     AND p.deleted = b'0'",
        "     AND p.tenant_id = s.tenant_id",
        "    WHERE p.id IS NULL",
        "    ) diff_product;",
        "    IF v_product_diff > 0 THEN",
        "      SIGNAL SQLSTATE '45000' SET MESSAGE_TEXT = 'SHOWROOM_SEED_BLOCKED: product master data no longer matches the initialization seed';",
        "    END IF;",
        "  END IF;",
        "",
        "  SELECT COUNT(*) INTO v_product_revision_rows FROM showroom_product_revision",
        "  WHERE tenant_id = v_seed_tenant_id AND deleted = b'0';",
        "  IF v_product_revision_rows > 0 THEN",
        "    SELECT COUNT(*) INTO v_product_revision_diff FROM (",
        "    SELECT r.id",
        "    FROM showroom_product_revision r",
        "    LEFT JOIN tmp_showroom_seed_product_revision s",
        "      ON r.id = s.id",
        "     AND r.product_id = s.product_id",
        "     AND r.revision_no = s.revision_no",
        "     AND r.status = s.status",
        "     AND IFNULL(r.name_cn, '') = IFNULL(s.name_cn, '')",
        "     AND IFNULL(r.name_en, '') = IFNULL(s.name_en, '')",
        "     AND IFNULL(r.owner_company_id, 0) = IFNULL(s.owner_company_id, 0)",
        "     AND IFNULL(r.product_owner_type, '') = IFNULL(s.product_owner_type, '')",
        "     AND IFNULL(r.lifecycle_stage, '') = IFNULL(s.lifecycle_stage, '')",
        "     AND IFNULL(r.registration_certificate, '') = IFNULL(s.registration_certificate, '')",
        "     AND IFNULL(r.indication_content, '') = IFNULL(s.indication_content, '')",
        "     AND r.tenant_id = s.tenant_id",
        "    WHERE r.tenant_id = v_seed_tenant_id AND r.deleted = b'0' AND s.id IS NULL",
        "    UNION ALL",
        "    SELECT s.id",
        "    FROM tmp_showroom_seed_product_revision s",
        "    LEFT JOIN showroom_product_revision r",
        "      ON r.id = s.id",
        "     AND r.deleted = b'0'",
        "     AND r.tenant_id = s.tenant_id",
        "    WHERE r.id IS NULL",
        "    ) diff_product_revision;",
        "    IF v_product_revision_diff > 0 THEN",
        "      SIGNAL SQLSTATE '45000' SET MESSAGE_TEXT = 'SHOWROOM_SEED_BLOCKED: product revisions no longer match the initialization seed';",
        "    END IF;",
        "  END IF;",
        "",
        "  SELECT COUNT(*) INTO v_hall_rows FROM showroom_hall",
        "  WHERE tenant_id = v_seed_tenant_id AND deleted = b'0';",
        "  IF v_hall_rows > 0 THEN",
        "    SELECT COUNT(*) INTO v_hall_diff FROM (",
        "    SELECT h.id",
        "    FROM showroom_hall h",
        "    LEFT JOIN tmp_showroom_seed_hall s",
        "      ON h.id = s.id",
        "     AND h.hall_code = s.hall_code",
        "     AND h.name = s.name",
        "     AND IFNULL(h.name_en, '') = IFNULL(s.name_en, '')",
        "     AND IFNULL(h.description, '') = IFNULL(s.description, '')",
        "     AND IFNULL(h.description_en, '') = IFNULL(s.description_en, '')",
        "     AND h.display_order = s.display_order",
        "     AND h.status = s.status",
        "     AND h.tenant_id = s.tenant_id",
        "    WHERE h.tenant_id = v_seed_tenant_id AND h.deleted = b'0' AND s.id IS NULL",
        "    UNION ALL",
        "    SELECT s.id",
        "    FROM tmp_showroom_seed_hall s",
        "    LEFT JOIN showroom_hall h",
        "      ON h.id = s.id",
        "     AND h.deleted = b'0'",
        "     AND h.tenant_id = s.tenant_id",
        "    WHERE h.id IS NULL",
        "    ) diff_hall;",
        "    IF v_hall_diff > 0 THEN",
        "      SIGNAL SQLSTATE '45000' SET MESSAGE_TEXT = 'SHOWROOM_SEED_BLOCKED: hall metadata no longer matches the initialization seed';",
        "    END IF;",
        "  END IF;",
        "",
        "  SELECT COUNT(*) INTO v_hall_product_rows FROM showroom_hall_product",
        "  WHERE tenant_id = v_seed_tenant_id AND deleted = b'0';",
        "  IF v_hall_product_rows > 0 THEN",
        "    SELECT COUNT(*) INTO v_hall_product_diff FROM (",
        "    SELECT hp.id",
        "    FROM showroom_hall_product hp",
        "    LEFT JOIN tmp_showroom_seed_hall_product s",
        "      ON hp.id = s.id",
        "     AND hp.hall_id = s.hall_id",
        "     AND hp.product_id = s.product_id",
        "     AND hp.display_order = s.display_order",
        "     AND hp.tenant_id = s.tenant_id",
        "    WHERE hp.tenant_id = v_seed_tenant_id AND hp.deleted = b'0' AND s.id IS NULL",
        "    UNION ALL",
        "    SELECT s.id",
        "    FROM tmp_showroom_seed_hall_product s",
        "    LEFT JOIN showroom_hall_product hp",
        "      ON hp.id = s.id",
        "     AND hp.deleted = b'0'",
        "     AND hp.tenant_id = s.tenant_id",
        "    WHERE hp.id IS NULL",
        "    ) diff_hall_product;",
        "    IF v_hall_product_diff > 0 THEN",
        "      SIGNAL SQLSTATE '45000' SET MESSAGE_TEXT = 'SHOWROOM_SEED_BLOCKED: hall-product mappings no longer match the initialization seed';",
        "    END IF;",
        "  END IF;",
        "",
        "  START TRANSACTION;",
        "    DELETE FROM showroom_hall_product WHERE tenant_id = v_seed_tenant_id;",
        "    DELETE FROM showroom_product_revision WHERE tenant_id = v_seed_tenant_id;",
        "    DELETE FROM showroom_hall WHERE tenant_id = v_seed_tenant_id;",
        "    DELETE FROM showroom_product WHERE tenant_id = v_seed_tenant_id;",
        "    DELETE FROM showroom_company WHERE tenant_id = v_seed_tenant_id;",
        "",
        "    INSERT INTO showroom_company (id, company_code, display_name, company_type, current_revision_id,",
        "        current_revision_no, incomplete_flag, status, creator, create_time, updater, update_time, deleted, tenant_id)",
        "    SELECT id, company_code, display_name, company_type, current_revision_id, current_revision_no,",
        "        incomplete_flag, status, creator, create_time, updater, update_time, deleted, tenant_id",
        "    FROM tmp_showroom_seed_company ORDER BY id;",
        "",
        "    INSERT INTO showroom_hall (id, hall_code, name, name_en, description, description_en, display_order, status, creator, create_time, updater, update_time, deleted, tenant_id)",
        "    SELECT id, hall_code, name, name_en, description, description_en, display_order, status, creator, create_time, updater, update_time, deleted, tenant_id",
        "    FROM tmp_showroom_seed_hall ORDER BY id;",
        "",
        "    INSERT INTO showroom_product (id, product_code, current_revision_id, current_revision_no, incomplete_flag, status, creator, create_time, updater, update_time, deleted, tenant_id)",
        "    SELECT id, product_code, current_revision_id, current_revision_no, incomplete_flag, status, creator, create_time, updater, update_time, deleted, tenant_id",
        "    FROM tmp_showroom_seed_product ORDER BY id;",
        "",
        "    INSERT INTO showroom_product_revision (id, product_id, revision_no, status, name_cn, name_en, owner_company_id, product_owner_type, lifecycle_stage, target_market, pipeline_layout, registration_certificate, indication_content, core_selling_points, model_specification, clinical_effect, fim_status, submitted_by, approved_by, published_at, creator, create_time, updater, update_time, deleted, tenant_id)",
        "    SELECT id, product_id, revision_no, status, name_cn, name_en, owner_company_id, product_owner_type, lifecycle_stage, target_market, pipeline_layout, registration_certificate, indication_content, core_selling_points, model_specification, clinical_effect, fim_status, submitted_by, approved_by, published_at, creator, create_time, updater, update_time, deleted, tenant_id",
        "    FROM tmp_showroom_seed_product_revision ORDER BY id;",
        "",
        "    INSERT INTO showroom_hall_product (id, hall_id, product_id, display_order, creator, create_time, updater, update_time, deleted, tenant_id)",
        "    SELECT id, hall_id, product_id, display_order, creator, create_time, updater, update_time, deleted, tenant_id",
        "    FROM tmp_showroom_seed_hall_product ORDER BY id;",
        "  COMMIT;",
        "END //",
        "DELIMITER ;",
        "",
        "CALL init_showroom_excel_seed();",
        "DROP PROCEDURE IF EXISTS init_showroom_excel_seed;",
        "",
    ]
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate showroom Excel initialization seed SQL.")
    parser.add_argument("--input", type=Path, default=default_input_path())
    parser.add_argument("--output", type=Path, default=default_output_path())
    args = parser.parse_args()

    companies, halls, products, empty_descriptions = parse_workbook(args.input)
    text = render_sql(companies, halls, products, empty_descriptions, args.input)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(text, encoding="utf-8")


if __name__ == "__main__":
    main()
