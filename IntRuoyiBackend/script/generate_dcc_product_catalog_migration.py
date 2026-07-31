from __future__ import annotations

import argparse
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXPECTED_HEADERS = [
    "产品类别I",
    "产品类别II",
    "产品序号",
    "产品",
    "产品编码",
    "注册证名称",
    "注册证号",
    "持证人",
    "注册地",
    "生效日期",
    "有效期至",
    "分类",
    "注册证信息链接",
    "产品状态（在研N/在售S/已取消C）",
    "备注",
]
SHEETS = [
    ("瑛泰产品（含璞慧、七木）", "瑛泰产品"),
]
COLUMNS = [
    "data_source",
    "original_row_no",
    "category_level1",
    "category_level2",
    "product_sequence",
    "product",
    "product_code",
    "registration_certificate_name",
    "registration_certificate_number",
    "certificate_holder",
    "registration_place",
    "effective_date",
    "expiry_date",
    "classification",
    "registration_info_link",
    "product_status",
    "remark",
    "creator",
    "updater",
]


def normalize_header(value: Any) -> str | None:
    text = normalize_text(value)
    if text is None:
        return None
    normalized = "".join(text.replace("　", "").split())
    if normalized == "注册证信息链接（国家药监局数据库）":
        return "注册证信息链接"
    return normalized


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).replace("\u00a0", " ").strip()
    return text or None


def read_rows(source: Path) -> list[list[Any]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    result: list[list[Any]] = []
    for sheet_name, data_source in SHEETS:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"missing sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        sheet.reset_dimensions()
        row_iterator = sheet.iter_rows(values_only=True)
        try:
            header_row = next(row_iterator)
        except StopIteration as exc:
            raise ValueError(f"empty sheet: {sheet_name}") from exc
        headers = [item for item in (normalize_header(value) for value in header_row) if item]
        if headers != EXPECTED_HEADERS:
            raise ValueError(f"invalid headers in {sheet_name}: {headers}")

        fill_down: list[str | None] = [None, None, None, None]
        for original_row_no, raw_row in enumerate(row_iterator, start=2):
            values = [normalize_text(raw_row[index] if index < len(raw_row) else None) for index in range(15)]
            if all(value is None for value in values):
                continue
            for index in range(4):
                if values[index] is not None:
                    fill_down[index] = values[index]
                else:
                    values[index] = fill_down[index]
            result.append([data_source, original_row_no, *values, "dcc-catalog-seed", "dcc-catalog-seed"])
    return result


def sql_literal(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, int):
        return str(value)
    return "'" + str(value).replace("\\", "\\\\").replace("'", "''") + "'"


def render_sql(rows: list[list[Any]], source: Path) -> str:
    value_lines = []
    for index, row in enumerate(rows):
        suffix = "," if index < len(rows) - 1 else ""
        value_lines.append("  (" + ", ".join(sql_literal(value) for value in row) + ")" + suffix)
    return "\n".join(
        [
            "-- release-migration: allowedEnvironments=test,backup,prod; dependsOn=20260513_dcc_base_schema; type=schema; riskLevel=medium",
            "-- DCC 产品目录数据库化：从确认版 Excel 一次性导入，运行时不再读取桌面文件。",
            f"-- Source workbook: {source}",
            "-- Rollback: export dcc_product_catalog if user changes exist, then DROP TABLE `dcc_product_catalog`.",
            "",
            "SET NAMES utf8mb4 COLLATE utf8mb4_unicode_ci;",
            "",
            "CREATE TABLE IF NOT EXISTS `dcc_product_catalog` (",
            "  `id` bigint NOT NULL AUTO_INCREMENT COMMENT '编号',",
            "  `data_source` varchar(64) NOT NULL COMMENT '数据来源',",
            "  `original_row_no` int NOT NULL COMMENT '稳定来源行号',",
            "  `category_level1` varchar(255) DEFAULT NULL COMMENT '产品类别 I',",
            "  `category_level2` varchar(255) DEFAULT NULL COMMENT '产品类别 II',",
            "  `product_sequence` varchar(64) DEFAULT NULL COMMENT '产品序号',",
            "  `product` varchar(512) NOT NULL COMMENT '产品',",
            "  `product_code` varchar(128) DEFAULT NULL COMMENT '产品编码',",
            "  `registration_certificate_name` varchar(512) DEFAULT NULL COMMENT '注册证名称',",
            "  `registration_certificate_number` varchar(255) DEFAULT NULL COMMENT '注册证号',",
            "  `certificate_holder` varchar(255) DEFAULT NULL COMMENT '持证人',",
            "  `registration_place` varchar(255) DEFAULT NULL COMMENT '注册地',",
            "  `effective_date` varchar(64) DEFAULT NULL COMMENT '生效日期',",
            "  `expiry_date` varchar(64) DEFAULT NULL COMMENT '有效期至',",
            "  `classification` varchar(128) DEFAULT NULL COMMENT '分类',",
            "  `registration_info_link` varchar(1024) DEFAULT NULL COMMENT '注册证信息链接',",
            "  `product_status` varchar(32) DEFAULT NULL COMMENT '产品状态',",
            "  `remark` varchar(1024) DEFAULT NULL COMMENT '备注',",
            "  `creator` varchar(64) DEFAULT '' COMMENT '创建者',",
            "  `create_time` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间',",
            "  `updater` varchar(64) DEFAULT '' COMMENT '更新者',",
            "  `update_time` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP COMMENT '更新时间',",
            "  `deleted` bit(1) NOT NULL DEFAULT b'0' COMMENT '是否删除',",
            "  PRIMARY KEY (`id`),",
            "  UNIQUE KEY `uk_dcc_product_catalog_source_row` (`data_source`, `original_row_no`),",
            "  KEY `idx_dcc_product_catalog_product` (`product`),",
            "  KEY `idx_dcc_product_catalog_product_code` (`product_code`),",
            "  KEY `idx_dcc_product_catalog_status` (`product_status`)",
            ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci COMMENT='DCC 产品目录';",
            "",
            "INSERT INTO `dcc_product_catalog` (",
            "  `" + "`, `".join(COLUMNS) + "`",
            ") VALUES",
            *value_lines,
            "ON DUPLICATE KEY UPDATE `id` = `id`;",
            "",
        ]
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    rows = read_rows(args.source)
    counts = {data_source: sum(1 for row in rows if row[0] == data_source) for _, data_source in SHEETS}
    if counts != {"瑛泰产品": 181}:
        raise ValueError(f"unexpected row counts: {counts}")
    args.output.write_text(render_sql(rows, args.source), encoding="utf-8", newline="\n")
    print(f"generated {len(rows)} rows: {counts}")


if __name__ == "__main__":
    main()
