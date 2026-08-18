import argparse
import hashlib
import json
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DEFAULT_SOURCE_WORKBOOK = Path(r"C:\Users\BJB110\Desktop\文档\1\医疗器械注册证信息表20260814.xlsx")
EXPECTED_SOURCE_SHA256 = "D42162DC354E8976CED450FA8A2BB00A2AB6099EDDF19AB907FEC3366EF94FF4"
MIGRATION_PREFLIGHT_PERMISSION = "dcc:registration-certificate:migration:preflight"
MIGRATION_ADMIN_PERMISSION = "system:migration:admin"
DOMESTIC_SHEET = "国内注册证"
EXCLUDED_SHEETS = ("CE证书", "FDA证书", "公司证件")
HEADERS = ["公司名称", "项目代码", "产品名称", "注册证号", "首次获证日期", "生效日期", "有效期至", "类别", "备注"]


class MigrationPreflightError(RuntimeError):
    pass


class MigrationPreflightPermissionError(PermissionError):
    pass


def assert_preflight_permission(permissions):
    permission_set = set(permissions or ())
    if MIGRATION_PREFLIGHT_PERMISSION not in permission_set and MIGRATION_ADMIN_PERMISSION not in permission_set:
        raise MigrationPreflightPermissionError("registration certificate migration preflight requires global admin permission")


def build_preflight_report(workbook_path=DEFAULT_SOURCE_WORKBOOK, permissions=None, expected_sha256=EXPECTED_SOURCE_SHA256):
    assert_preflight_permission(permissions or ())
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise MigrationPreflightError(f"source workbook not found: {workbook_path}")
    source_sha256 = sha256_file(workbook_path)
    if expected_sha256 and source_sha256.upper() != expected_sha256.upper():
        raise MigrationPreflightError(
            f"source workbook sha256 mismatch: expected {expected_sha256.upper()}, actual {source_sha256.upper()}"
        )

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if DOMESTIC_SHEET not in workbook.sheetnames:
            raise MigrationPreflightError(f"workbook missing required sheet: {DOMESTIC_SHEET}")
        domestic = workbook[DOMESTIC_SHEET]
        headers = [text(cell.value) for cell in next(domestic.iter_rows(min_row=1, max_row=1))]
        if headers[:len(HEADERS)] != HEADERS or len([h for h in headers if h]) != len(HEADERS):
            raise MigrationPreflightError(f"domestic sheet header mismatch: {headers}")

        rows = []
        company_names = set()
        missing_project_count = 0
        for row_number, cells in enumerate(domestic.iter_rows(min_row=2, max_col=len(HEADERS)), start=2):
            raw = {header: text(cells[index].value) for index, header in enumerate(HEADERS)}
            if not any(raw.values()):
                continue
            company_name = normalized_text(raw["公司名称"])
            project_code = normalized_text(raw["项目代码"])
            product_name = normalized_text(raw["产品名称"])
            certificate_no = normalized_text(raw["注册证号"])
            classification = normalized_text(raw["类别"])
            remark = normalized_text(raw["备注"])
            first_date = normalize_date(raw["首次获证日期"])
            effective_date = normalize_date(raw["生效日期"])
            expiry_date = normalize_date(raw["有效期至"])
            if company_name:
                company_names.add(company_name)
            if project_code is None:
                missing_project_count += 1
            rows.append(build_review_row(
                row_number=row_number,
                cells=cells,
                raw=raw,
                normalized={
                    "owner_company_name": company_name,
                    "project_code": project_code,
                    "product_name": product_name,
                    "certificate_no": certificate_no,
                    "first_obtained_date": first_date,
                    "effective_date": effective_date,
                    "expiry_date": expiry_date,
                    "classification": classification,
                    "remark": remark,
                    "approval_date": None,
                    "registrant_name": None,
                    "production_relation": None,
                    "attachment_evidence": None,
                },
            ))

        excluded_counts = {sheet: count_data_rows(workbook[sheet]) if sheet in workbook.sheetnames else 0
                           for sheet in EXCLUDED_SHEETS}
        sheet_import_counts = {DOMESTIC_SHEET: len(rows)}
        sheet_import_counts.update({sheet: 0 for sheet in EXCLUDED_SHEETS})
        return {
            "source_path": str(workbook_path),
            "source_sha256": source_sha256.upper(),
            "domestic_row_count": len(rows),
            "company_count": len(company_names),
            "missing_project_code_count": missing_project_count,
            "sheet_import_counts": sheet_import_counts,
            "excluded_sheet_row_counts": excluded_counts,
            "rows": rows,
        }
    finally:
        workbook.close()


def sha256_file(path):
    hasher = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def normalized_text(value):
    value = text(value)
    return value if value else None


def normalize_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    value = text(value)
    if not value:
        return None
    match = re.fullmatch(r"(\d{4})[./-](\d{1,2})[./-](\d{1,2})", value)
    if not match:
        return None
    year, month, day = (int(part) for part in match.groups())
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return None


def build_review_row(row_number, cells, raw, normalized):
    evidence = {}
    for index, header in enumerate(HEADERS):
        raw_value = raw[header]
        evidence[field_key(header)] = {
            "source": f"{DOMESTIC_SHEET}!{get_column_letter(index + 1)}{row_number}",
            "raw": raw_value,
            "status": "SOURCE_VALUE" if raw_value else "BLANK_SOURCE",
        }
    for missing_field in ("approval_date", "registrant_name", "production_relation", "attachment_evidence"):
        evidence[missing_field] = {
            "source": None,
            "raw": None,
            "status": "MISSING_FORMAL_EVIDENCE",
        }

    review_reasons = [
        "missing_owner_company_mapping",
        "missing_product_mapping",
        "missing_approval_date",
        "missing_registrant_name",
        "missing_production_relation",
        "missing_attachment_evidence",
        "classification_needs_review",
    ]
    if normalized["project_code"] is None:
        review_reasons.append("missing_project_code")
    else:
        review_reasons.append("missing_project_code_mapping")
    for key in ("first_obtained_date", "effective_date", "expiry_date"):
        if normalized[key] is None:
            review_reasons.append(f"invalid_{key}")

    return {
        "source_sheet": DOMESTIC_SHEET,
        "source_row": row_number,
        "raw": raw,
        "normalized": normalized,
        "evidence": evidence,
        "review_status": "NEEDS_REVIEW",
        "review_reasons": sorted(set(review_reasons)),
    }


def field_key(header):
    return {
        "公司名称": "owner_company_name",
        "项目代码": "project_code",
        "产品名称": "product_name",
        "注册证号": "certificate_no",
        "首次获证日期": "first_obtained_date",
        "生效日期": "effective_date",
        "有效期至": "expiry_date",
        "类别": "classification",
        "备注": "remark",
    }[header]


def count_data_rows(sheet):
    return max(sheet.max_row - 1, 0)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", default=str(DEFAULT_SOURCE_WORKBOOK))
    parser.add_argument("--permission", action="append", default=[MIGRATION_PREFLIGHT_PERMISSION])
    parser.add_argument("--expected-sha256", default=EXPECTED_SOURCE_SHA256)
    parser.add_argument("--output-json")
    args = parser.parse_args()

    report = build_preflight_report(
        args.workbook,
        permissions=set(args.permission or ()),
        expected_sha256=args.expected_sha256 or None,
    )
    payload = json.dumps(report, ensure_ascii=False, indent=2)
    if args.output_json:
        output_path = Path(args.output_json)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(payload + "\n", encoding="utf-8")
    print(payload)


if __name__ == "__main__":
    main()
