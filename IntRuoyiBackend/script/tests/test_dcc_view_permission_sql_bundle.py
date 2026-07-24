import json
import subprocess
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[2]
SCRIPT = ROOT / "script" / "dcc_view_permission_sql_bundle.py"


def write_matrix_workbook(path, classification_confirm="", role_confirm=""):
    wb = Workbook()
    classification = wb.active
    classification.title = "文件归类待确认"
    classification.append(["deleted", "file_id", "tenant_id", "manual_confirm_category_code"])
    classification.append(["0", "100", "1", classification_confirm])
    roles = wb.create_sheet("主管角色候选")
    roles.append(["role_code", "candidate_user_id", "manual_confirm"])
    roles.append(["dcc_matrix_qc_lead", "200", role_confirm])
    wb.save(path)


def write_product_group_workbook(path, manual_confirm=""):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "候选明细"
    headers = [
        "tenant_id",
        "group_code",
        "group_name",
        "dept_id",
        "user_id",
        "product_master_id",
        "manual_confirm",
    ]
    sheet.append(headers)
    sheet.append(["1", "np-alpha", "新品A组", "136", "200", "300", manual_confirm])
    wb.save(path)


def run_bundle(tmp_path, matrix_confirm="", role_confirm="", product_confirm=""):
    matrix = tmp_path / "matrix.xlsx"
    product = tmp_path / "product.xlsx"
    output_dir = tmp_path / "bundle"
    write_matrix_workbook(matrix, classification_confirm=matrix_confirm, role_confirm=role_confirm)
    write_product_group_workbook(product, manual_confirm=product_confirm)
    cp = subprocess.run(
        [
            "python", "-X", "utf8", str(SCRIPT),
            "--matrix-workbook", str(matrix),
            "--product-group-workbook", str(product),
            "--output-dir", str(output_dir),
        ],
        cwd=ROOT,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )
    return cp, output_dir


def test_bundle_fails_and_removes_outputs_when_gate_fails(tmp_path):
    cp, output_dir = run_bundle(tmp_path)

    assert cp.returncode != 0
    assert "matrix workbook has no confirmed" in cp.stderr
    assert not output_dir.exists()


def test_bundle_generates_ordered_sql_files_and_manifest(tmp_path):
    cp, output_dir = run_bundle(
        tmp_path,
        matrix_confirm="DCC_FVM_DMR_001",
        role_confirm="确认",
        product_confirm="确认",
    )

    assert cp.returncode == 0, cp.stdout + cp.stderr
    manifest = json.loads((output_dir / "manifest.json").read_text(encoding="utf-8"))
    assert manifest["ready"] is True
    assert manifest["confirmedFiles"] == 1
    assert manifest["confirmedRoles"] == 1
    assert manifest["confirmedProductGroupRows"] == 1
    assert manifest["executionOrder"] == [
        "01-dcc-matrix-confirmed.sql",
        "02-dcc-product-group-confirmed.sql",
    ]
    assert set(manifest["inputFiles"]) == {
        "matrixClassificationCsv",
        "matrixRoleCsv",
        "productGroupCsv",
    }
    assert (output_dir / "confirmation-gate-result.json").exists()
    assert (output_dir / "inputs" / "matrix-file-classification.csv").exists()
    assert (output_dir / "inputs" / "matrix-role-members.csv").exists()
    assert (output_dir / "inputs" / "product-group-bindings.csv").exists()
    assert "UPDATE dcc_controlled_file" in (output_dir / "01-dcc-matrix-confirmed.sql").read_text(encoding="utf-8")
    assert "dcc_product_visibility_group" in (output_dir / "02-dcc-product-group-confirmed.sql").read_text(encoding="utf-8")
