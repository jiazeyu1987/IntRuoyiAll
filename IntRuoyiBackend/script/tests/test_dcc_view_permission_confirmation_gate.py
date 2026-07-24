import json
import subprocess
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[2]
SCRIPT = ROOT / "script" / "dcc_view_permission_confirmation_gate.py"


def write_matrix_workbook(path, classification_confirm="", role_confirm=""):
    wb = Workbook()
    classification = wb.active
    classification.title = "文件归类待确认"
    classification.append(["deleted", "file_id", "tenant_id", "manual_confirm_category_code"])
    classification.append(["0", "100", "1", classification_confirm])
    roles = wb.create_sheet("主管角色候选")
    roles.append(["role_code", "candidate_user_id", "manual_confirm"])
    roles.append(["dcc_matrix_qc_lead", "200", role_confirm])
    wb.save(path)


def write_product_group_workbook(path, manual_confirm=""):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "候选明细"
    headers = [
        "tenant_id",
        "group_code",
        "group_name",
        "dept_id",
        "user_id",
        "product_master_id",
        "manual_confirm",
    ]
    sheet.append(headers)
    sheet.append(["1", "np-alpha", "新品A组", "136", "200", "300", manual_confirm])
    wb.save(path)


def run_gate(tmp_path, matrix_confirm="", role_confirm="", product_confirm=""):
    matrix = tmp_path / "matrix.xlsx"
    product = tmp_path / "product.xlsx"
    output = tmp_path / "gate.json"
    write_matrix_workbook(matrix, classification_confirm=matrix_confirm, role_confirm=role_confirm)
    write_product_group_workbook(product, manual_confirm=product_confirm)
    cp = subprocess.run(
        [
            "python", "-X", "utf8", str(SCRIPT),
            "--matrix-workbook", str(matrix),
            "--product-group-workbook", str(product),
            "--output-json", str(output),
        ],
        cwd=ROOT,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )
    payload = json.loads(output.read_text(encoding="utf-8")) if output.exists() else None
    return cp, payload


def test_gate_fails_when_both_workbooks_have_no_confirmed_rows(tmp_path):
    cp, payload = run_gate(tmp_path)

    assert cp.returncode != 0
    assert payload["ready"] is False
    assert "matrix workbook has no confirmed" in " ".join(payload["reasons"])
    assert "product group workbook has no confirmed rows" in " ".join(payload["reasons"])


def test_gate_passes_when_workbooks_have_confirmed_rows(tmp_path):
    cp, payload = run_gate(tmp_path, matrix_confirm="DCC_FVM_DMR_001", role_confirm="确认", product_confirm="确认")

    assert cp.returncode == 0, cp.stdout + cp.stderr
    assert payload["ready"] is True
    assert payload["matrix"]["confirmedFiles"] == 1
    assert payload["matrix"]["confirmedRoles"] == 1
    assert payload["productGroup"]["confirmedProductGroupRows"] == 1


def test_gate_reports_invalid_matrix_confirmation(tmp_path):
    cp, payload = run_gate(tmp_path, matrix_confirm="bad-category", product_confirm="确认")

    assert cp.returncode != 0
    assert payload["ready"] is False
    assert "invalid matrix category code" in " ".join(payload["reasons"])
