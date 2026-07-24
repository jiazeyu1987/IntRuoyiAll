import argparse
import csv
import sys
from pathlib import Path

from openpyxl import load_workbook


TRUE_VALUES = {"1", "true", "yes", "y", "是", "确认", "confirmed"}
SHEET_NAME = "候选明细"
OUTPUT_FIELDS = [
    "tenant_id",
    "group_code",
    "group_name",
    "dept_id",
    "user_id",
    "product_master_id",
    "manual_confirm",
]
REQUIRED_COLUMNS = set(OUTPUT_FIELDS)


def normalize_bool(value):
    return str(value or "").strip().lower() in TRUE_VALUES


def normalize_cell(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def load_rows(input_xlsx):
    workbook = load_workbook(input_xlsx, data_only=True, read_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Missing worksheet: {SHEET_NAME}")
    sheet = workbook[SHEET_NAME]
    header_values = [normalize_cell(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_map = {name: index for index, name in enumerate(header_values) if name}
    missing = sorted(REQUIRED_COLUMNS - set(header_map))
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    rows = []
    for row_index, cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        record = {field: normalize_cell(cells[header_map[field]].value) for field in OUTPUT_FIELDS}
        if not any(record.values()):
            continue
        if not normalize_bool(record["manual_confirm"]):
            continue
        missing_values = [field for field in OUTPUT_FIELDS if not record[field]]
        if missing_values:
            raise ValueError(f"Confirmed row {row_index} missing required values: {', '.join(missing_values)}")
        rows.append(record)
    return rows


def write_csv(path, rows):
    with Path(path).open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-xlsx", required=True)
    parser.add_argument("--output-csv", required=True)
    args = parser.parse_args()

    try:
        rows = load_rows(args.input_xlsx)
        if not rows:
            raise ValueError("No confirmed rows found")
        write_csv(args.output_csv, rows)
        print(f"confirmed_rows={len(rows)}")
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
