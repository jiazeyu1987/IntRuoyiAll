import argparse
import csv
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SQL_SCRIPT = ROOT / "script" / "dcc_matrix_confirmed_sql_generator.py"

CLASSIFICATION_SHEET = "文件归类待确认"
ROLE_SHEET = "主管角色候选"

CLASSIFICATION_FIELDS = [
    "file_id",
    "tenant_id",
    "deleted",
    "manual_confirm_category_code",
]

ROLE_FIELDS = [
    "role_code",
    "candidate_user_id",
    "manual_confirm",
]


def row_to_dict(headers, row):
    return {
        str(header).strip(): ("" if value is None else value)
        for header, value in zip(headers, row)
        if header is not None and str(header).strip()
    }


def read_sheet_rows(workbook, sheet_name, required_fields):
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Workbook missing required sheet: {sheet_name}")
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Sheet has no header row: {sheet_name}")
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    missing = [field for field in required_fields if field not in headers]
    if missing:
        raise ValueError(f"Sheet {sheet_name} missing required headers: {', '.join(missing)}")
    return [row_to_dict(headers, row) for row in rows[1:]]


def write_csv(path, fields, rows):
    with Path(path).open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def run_command(args):
    cp = subprocess.run(
        args,
        cwd=ROOT,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )
    if cp.returncode != 0:
        raise RuntimeError((cp.stderr or cp.stdout).strip())
    return cp.stdout.strip()


def export_workbook(input_xlsx, classification_csv, role_csv):
    workbook = load_workbook(input_xlsx, read_only=False, data_only=True)
    classification_rows = read_sheet_rows(workbook, CLASSIFICATION_SHEET, CLASSIFICATION_FIELDS)
    role_rows = read_sheet_rows(workbook, ROLE_SHEET, ROLE_FIELDS)
    write_csv(classification_csv, CLASSIFICATION_FIELDS, classification_rows)
    write_csv(role_csv, ROLE_FIELDS, role_rows)
    return len(classification_rows), len(role_rows)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-xlsx", required=True)
    parser.add_argument("--output-sql", required=True)
    parser.add_argument("--keep-classification-csv")
    parser.add_argument("--keep-role-csv")
    args = parser.parse_args()

    output_sql = Path(args.output_sql)
    output_sql.parent.mkdir(parents=True, exist_ok=True)
    temp_dir = None
    try:
        if bool(args.keep_classification_csv) != bool(args.keep_role_csv):
            raise ValueError("--keep-classification-csv and --keep-role-csv must be used together")
        if args.keep_classification_csv:
            classification_csv = Path(args.keep_classification_csv)
            role_csv = Path(args.keep_role_csv)
            classification_csv.parent.mkdir(parents=True, exist_ok=True)
            role_csv.parent.mkdir(parents=True, exist_ok=True)
        else:
            temp_dir = tempfile.TemporaryDirectory()
            classification_csv = Path(temp_dir.name) / "confirmed-file-classification.csv"
            role_csv = Path(temp_dir.name) / "confirmed-role-members.csv"

        classification_count, role_count = export_workbook(args.input_xlsx, classification_csv, role_csv)
        sql_stdout = run_command([
            sys.executable, "-X", "utf8", str(SQL_SCRIPT),
            "--classification-csv", str(classification_csv),
            "--role-csv", str(role_csv),
            "--output-sql", str(output_sql),
        ])
        print(f"classification_rows={classification_count}")
        print(f"role_rows={role_count}")
        if sql_stdout:
            print(sql_stdout)
    except Exception as exc:
        if output_sql.exists():
            output_sql.unlink()
        print(str(exc), file=sys.stderr)
        return 1
    finally:
        if temp_dir is not None:
            temp_dir.cleanup()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
